from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores=[
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
]
# 기존 성적 데이터 넣기
for s in scores:
    ws.append(s)

wrong_quize2 = ws.iter_cols(min_col=4,max_col=4,min_row=2)

#모든 퀴즈2의 점수 10으로 바꾸기
for quize_col in wrong_quize2:
    for quize_socore in quize_col:
        quize_socore.value = 10
        
# 총점 넣기
ws.cell(row=1, column=ws.max_column+1).value = "총점"
ws.cell(row=1, column=ws.max_column+1).value = "성적"
for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10 #총점
    ws.cell(row=idx, column=8).value = f"=SUM(B{idx}:G{idx})"
    
    #성적
    grade = None
    if score[1] < 5:
        grade = "F"
    elif sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"

    ws.cell(row=idx, column=9).value = grade


wb.save("scores.xlsx")
